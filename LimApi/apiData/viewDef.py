import copy
import datetime
import json
import os
import time
from urllib.parse import urlencode

import requests
from django.db.models import Max, F
from django.db.models.functions import JSONObject
from openpyxl import load_workbook
from requests import ReadTimeout
from rest_framework.response import Response

from apiData.models import ApiData, ApiCaseStep, ApiCase, ApiForeachStep
from comMethod.comDef import get_proj_envir_db_data, db_connect, execute_sql_func, \
    close_db_con, json_dumps, JSONEncoder, MyThread, json_loads, format_parm_type_v
from comMethod.constant import USER_API, VAR_PARAM, HEADER_PARAM, HOST_PARAM, RUNNING, SUCCESS, FAILED, DISABLED, \
    INTERRUPT, SKIP, API_CASE, API_FOREACH, TABLE_MODE, STRING, DIY_CFG, JSON_MODE, PY_TO_CONF_TYPE, CODE_MODE, \
    OBJECT, FAILED_STOP, WAITING, PRO_CFG, FORM_MODE, EQUAL, API_VAR, NOT_EQUAL, \
    CONTAIN, NOT_CONTAIN, TEXT_MODE, API, FORM_FILE_TYPE
from comMethod.diyException import DiyBaseException, NotFoundFileError
from comMethod.paramsDef import parse_param_value, run_params_code, parse_temp_params, get_parm_v_by_temp
from project.models import ProjectEnvirData
from user.models import UserCfg, UserTempParams


def create_api(req_data, default_params):
    """
    创建自定义Api用例基础数据
    """
    # 因为唯一性可能会报错
    api = ApiData.objects.create(
        name=req_data['name'], path=req_data['path'], method=req_data['method'], default_params=default_params,
        timeout=req_data.get('timeout'), project_id=req_data['project_id'], source=USER_API,
        module_id=req_data['module_id'])
    return api.id


def update_api(req_data, default_params, api_id, source):
    """
    更新自定义Api用例基础数据
    """

    update_fields = {'timeout': req_data.get('timeout'), 'default_params': default_params}
    if source == USER_API:
        update_fields.update({field: req_data.get(field) for field in ('name', 'path', 'method')})
        project_id, module_id = req_data.get('project_id'), req_data.get('module_id')
        if project_id:
            update_fields['project_id'] = project_id
        if module_id:
            update_fields['module_id'] = module_id
    ApiData.objects.filter(id=api_id).update(**update_fields)


def save_api(req_data, api_id, source):
    """
    创建Api的请求数据
    当在创建用例的同时创建接口时，需要传递 is_case=true/1,这样才能创建用例测试数据。
    """
    param_fields = []
    for key in ('header', 'query', 'body', 'expect', 'output'):
        param_fields.extend((f'{key}_source', f'{key}_mode'))
    param_fields.extend(('host', 'host_type'))
    default_params = {key: req_data.get(key) for key in param_fields}
    if api_id:
        update_api(req_data, default_params, api_id, source)
    else:
        api_id = create_api(req_data, default_params)
    return api_id


class ApiCasesActuator:
    """
    接口用例执行器
    user_id：执行计划的用户，传递了user_id时才会进行中断判断(调试时不会传递)
    temp_params:默认参数
    """

    def __init__(self, user_id, trigger_case_id=None, cfg_data=None, temp_params=None):
        self.user_id = user_id
        self.timeout = 60  # 默认接口超时时间
        self.base_params_source = {'user_id': self.user_id, 'case_id': trigger_case_id}
        user_cfg = UserCfg.objects.filter(user_id=user_id).values().first() or {
            'envir_id': 1, 'failed_stop': True, 'only_failed_log': False}
        cfg_data = {**user_cfg, **cfg_data} if cfg_data else user_cfg
        self.envir = str(cfg_data['envir_id'])
        self.failed_stop = cfg_data['failed_stop']
        self.only_failed_log = cfg_data['only_failed_log']
        self.status = SUCCESS
        self.cascader_error = False
        temp_params = temp_params or UserTempParams.objects.filter(user_id=user_id).values()
        params = parse_temp_params(temp_params)
        self.default_header, self.default_var, self.default_host = (params[key] for key in ('header', 'var', 'host'))
        self.params_source = params['params_source']
        self.user_id = user_id  # 执行计划的用户，传递了user_id时才会进行中断判断
        self.status = RUNNING  # 初始化执行状态为执行中
        self.api_data = {}  # 执行过的接口信息会存在这，避免频繁查库，示例:{id:{'path':/xx,'method':'GET','timeout':10}}
        self.api_process = ''

    @staticmethod
    def clear_upload_files(upload_files_list):
        """
        清除请求中上传的临时文件
        """
        for file in upload_files_list:
            file['file'] and file['file'].close()
            os.remove(file['name'])

    def parse_api_step_output(self, params, prefix_label, step_name, response, i):
        """
        处理api步骤的输出
        """
        out_data = {}
        if output := params.get('output_source'):
            self.api_process = '【输出参数】'
            if params['output_mode'] == TABLE_MODE:
                for out in output:
                    out_v, out_name = out['value'], str(parse_param_value(out['name'], self.default_var, i)).strip('?')
                    is_assert = not out['name'].endswith('?')
                    value_location_list = [parse_param_value(var, self.default_var, i) for var in out_v.split('.')]
                    try:
                        res = get_parm_v_by_temp(value_location_list, response)
                    except IndexError as e:
                        if is_assert:
                            return {'status': FAILED, 'results': str(e)}
                        else:
                            res = False
                    if not res:
                        if is_assert:
                            return {'status': FAILED, 'results': out_v + ':' + '未在响应结果中找到！'}
                    else:
                        res_v = res['value']
                        self.default_var[out_name] = res_v
                        out_data[out_name] = res_v
                        self.params_source[VAR_PARAM][out_name] = {
                            'name': out_name, 'value': res_v, 'step_name': prefix_label + step_name,
                            'type': VAR_PARAM, 'param_type_id': PY_TO_CONF_TYPE.get(str(type(res_v)), STRING),
                            **self.base_params_source}
            else:  # 代码模式
                res = run_params_code(output, response, self.default_var, i)
                if isinstance(res, dict):
                    self.default_var.update(res)
                    out_data = res
                    for name in res.keys():
                        self.params_source[VAR_PARAM][name] = {
                            'name': name, 'value': res[name], 'step_name': prefix_label + step_name,
                            'type': VAR_PARAM,
                            'param_type_id': PY_TO_CONF_TYPE.get(str(type(res[name])), STRING),
                            **self.base_params_source}
                else:
                    return {'status': FAILED, 'results': '返回数据格式不符合要求！'}
        return {'status': SUCCESS, 'out_data': out_data}

    def parse_api_step_expect(self, params, response, i):
        """
        处理api步骤的期望
        """
        old_default_var = copy.deepcopy(self.default_var)
        if expect := params.get('expect_source'):
            self.api_process = '【预期结果】'
            if params['expect_mode'] == TABLE_MODE:
                for ext in expect:
                    ext_name, ext_v = ext['name'], parse_param_value(ext['value'], old_default_var, i)
                    ext_v_type = ext.get('type', {}).get('type') or STRING
                    ext_v = format_parm_type_v(ext_v, ext_v_type)
                    rule = ext.get('rule', EQUAL)
                    # 使用str(parse_param_value(var, var_dict))是确保键为数字时能够转换为字符串数字进行匹配
                    ext_name_list = [str(parse_param_value(name, old_default_var, i)) for name in ext_name.split('.')]
                    res = get_parm_v_by_temp(ext_name_list, response)
                    if res:
                        res_v = res['value']
                        # res_v = str(res['value']) if ext_v_type == STRING else json_loads(res['value'])
                        print('ex', res_v, rule, type(ext_v))
                        if rule == EQUAL and res_v != ext_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 与响应结果不相等！响应结果值为:' + str(res_v) + ';\n'}
                        elif rule == NOT_EQUAL and res_v == ext_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 与响应结果相等【期望不相等】！响应结果值为:' + str(res_v) + ';\n'}
                        elif rule == CONTAIN and ext_v not in res_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 响应结果不包含该期望！响应结果值为:' + str(res_v) + ';\n'}
                        elif rule == NOT_CONTAIN and ext_v in res_v:
                            return {'status': FAILED, 'results': ext_name + '的期望值:' + str(
                                ext_v) + ' 响应结果包含该期望【期望不包含】！响应结果值为:' + str(res_v) + ';\n'}
                    else:
                        return {'status': FAILED, 'results': '未在响应中找到字段：' + ext_name + '\n'}
            else:  # 代码模式
                res = run_params_code(expect, old_default_var, i, response)
                if res is not None:
                    if isinstance(res, tuple) and res[0] is False:
                        return {'status': FAILED, 'results': res[1]}
                    elif res is False:
                        return {'status': FAILED, 'results': '不符合预期！'}
        return {'status': SUCCESS}

    def parse_excel_var_params(self, file_name):
        """
        解析form-data中excel存在的变量
        """
        workbook = load_workbook(file_name)
        sheets = workbook.sheetnames
        for sheet in sheets:
            ws = workbook[sheet]
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row, col)
                    cell.value = parse_param_value(cell.value, self.default_var)
        workbook.save(file_name)

    def parse_form_data_params(self, body, files_list):
        """
        解析form-data参数
        """
        req_data, body_log = {}, {}
        for key in body:
            field_data = body[key]
            if isinstance(field_data, dict) and 'type' in field_data:
                if field_data['type'] == FORM_FILE_TYPE:
                    file_name, file_url = field_data['name'], field_data['value']
                    print('ff', file_url)
                    r = requests.get(file_url)
                    if r.status_code == 404:
                        raise NotFoundFileError('未找到上传的文件：' + file_name)
                    with open(file_name, 'wb') as f:
                        f.write(r.content)
                    file_type = os.path.splitext(file_name)[-1]
                    if file_type in ['.xlsx'] and '有变量' in file_name:
                        try:
                            self.parse_excel_var_params(file_name)
                        except Exception as e:
                            files_list.append({'name': file_name, 'file': None})
                            raise Exception('导入报错：' + str(e))
                    file = open(file_name, 'rb')
                    req_data[key] = (file_name, file)
                    body_log[key] = file_name
                    files_list.append({'name': file_name, 'file': file})
                else:
                    body_log[key] = field_data['value']
                    req_data[key] = (None, field_data['value'])
            else:
                body_log = '不是有效的form-data参数！'
                break
        return req_data, body_log

    def api(self, step, prefix_label, i=0):
        """
          执行类型为接口的步骤
        """
        upload_files_list = []
        params = step['params']
        if host := params.get('host') or '':
            if params.get('host_type') == PRO_CFG:
                pro_data = ProjectEnvirData.objects.filter(
                    project_id=host, envir_id=self.envir).values_list('data', flat=True).first() or {}
                host = pro_data.get('host') or ''
        elif self.default_host:
            host = self.default_host
        # 参数中包含了case_id则走库里面取接口信息，不然则使用传递的（调试时才会无case_id）
        if api_id := params.get('api_id'):
            api_base = self.api_data.get(api_id)
            if not api_base:
                api_base = ApiData.objects.filter(id=params['api_id']).annotate(
                    api_base=JSONObject(path=F('path'), method=F('method'), timeout=F('timeout'))).values_list(
                    'api_base', flat=True).first()
                self.api_data[api_id] = api_base
            params.update(api_base)
        url, method, timeout = host + params['path'], params['method'], params.get('timeout') or self.timeout
        req_log = {'url': url, 'method': method, 'response': '无响应结果', 'res_header': '无响应头'}
        res_status, results = FAILED, ''
        try:
            self.api_process = '【Header(请求头)】'
            if header_source := params.get('header_source'):
                header = self.parse_source_params(header_source, params['header_mode'], i)
                header = {str(key).lower(): str(header[key]) for key in header}  # header的key全部转换为小写
                if not header.get('content-type'):
                    header['content-type'] = 'application/json'
                # 只有没有默认请求头时才将自定义的请求头设置为默认请求头，如果使用了全局参数且有默认请求头，则永远不会替换
                if not self.default_header:
                    self.default_header = copy.deepcopy(header)
            else:
                header = copy.deepcopy(self.default_header) or {'content-type': 'application/json'}
            self.api_process = '【query(url参数)】'
            query = self.parse_source_params(params.get('query_source'), params['query_mode'], i)
            self.api_process = '【Body(请求体)】'
            if params['body_mode'] != FORM_MODE:
                body = self.parse_source_params(params.get('body_source'), params['body_mode'], i)
            else:
                body = self.parse_source_params(
                    params.get('body_source'), params['body_mode'], i, file_list=upload_files_list)
            req_params = {'url': url, 'headers': header, 'params': query, 'method': method.lower(), 'timeout': timeout}
            req_log.update({'header': copy.deepcopy(header), 'body': body})
            content_type = header['content-type']
            if params['body_mode'] != FORM_MODE:
                if 'application/json' in content_type:
                    req_params['data'] = json_dumps(body).encode() if not isinstance(body, str) else body.encode(
                        'utf-8')
                elif 'text/html' in content_type:
                    req_params['data'] = body.encode('utf-8') if isinstance(body, str) else ''
                elif 'urlencoded' in content_type or 'text/plain' in content_type:
                    if not isinstance(body, dict):
                        req_params['data'] = body
                    else:
                        req_data = {k: json.dumps(body[k], ensure_ascii=False, separators=(',', ':')) if isinstance(
                            body[k], dict) else body[k] for k in body}
                        urlencode_v = urlencode(req_data).replace('+', '%20')
                        req_params['data'] = urlencode_v
            else:
                header.pop('content-type', None)
                req_log['header']['content-type'] = 'multipart/form-data'
                req_params['files'], req_log['body'] = body
            try:
                r = requests.request(**req_params)
            except IndexError as e:
                req_log['results'] = results = self.api_process + str(e)
            except KeyError as e:
                req_log['results'] = results = self.api_process + '未找到key：' + str(e)
            except (requests.exceptions.ConnectionError, ReadTimeout):
                req_log['response'] = results = '请求超时！'
            except requests.exceptions.InvalidSchema:
                req_log['results'] = results = '无效的请求地址！'
            except requests.exceptions.MissingSchema:
                req_log['results'] = results = '请求地址不能为空！'
            except Exception as e:
                req_log['results'] = results = str(e)
            else:
                spend_time = float('%.2f' % r.elapsed.total_seconds())
                try:
                    response = r.json()
                except IndexError as e:
                    response = r.text
                    if r.status_code == 404:
                        results = '请求地址不存在！'
                else:
                    out_res = self.parse_api_step_output(
                        params, prefix_label, step.get('step_name', '未命名步骤'), response, i)
                    res_status, results = out_res['status'], out_res.get('results')
                    if res_status == FAILED:
                        results = self.api_process + results
                    elif out_data := out_res.get('out_data'):
                        req_log['output'] = out_data
                    ext_res = self.parse_api_step_expect(params, response, i)
                    if res_status != FAILED:
                        res_status = ext_res['status']
                    if ext_res['status'] == FAILED:
                        results = self.api_process + ext_res.get('results', '')
                req_log.update({'url': r.url, 'res_header': dict(r.headers), 'response': response,
                                'spend_time': spend_time, 'results': results})
        except Exception as e:
            print('api_error', str(e), e.__traceback__.tb_lineno)
            req_log['results'] = results = self.api_process + str(e)
        self.clear_upload_files(upload_files_list)
        return {'status': res_status, 'results': {'msg': results, 'request_log': req_log}}

    def header(self, step, prefix_label, i=0):
        """
         执行类型为请求头的步骤
        """
        if params := step['params']:
            header = self.parse_source_params(params, i=i)
            self.default_header = header
            for name in header.keys():
                self.params_source[HEADER_PARAM][name] = {
                    'name': name, 'value': header[name], 'step_name': prefix_label + step['step_name'],
                    'type': HEADER_PARAM, 'param_type_id': STRING, **self.base_params_source}

    def host(self, step, prefix_label, i=0):
        """
        执行类型为域名的步骤
        """
        params = step['params']
        if params['host_type'] == DIY_CFG:
            self.default_host = params['value']
        else:
            data = ProjectEnvirData.objects.filter(project_id=params['value']).values_list(
                'data', flat=True).first() or {}
            self.default_host = data.get('host')
            if not self.default_host:
                return {'status': FAILED, 'results': '该项目没有配置请求地址！'}
        self.params_source[HOST_PARAM]['请求地址'] = {
            'name': '请求地址', 'value': self.default_host, 'type': HOST_PARAM,
            'step_name': prefix_label + step['step_name'], 'param_type_id': STRING, **self.base_params_source}

    def case(self, step, prefix_label='', cascader_level=1, i=0):
        """
        执行类型为用例
        """

        if cascader_level > 10:
            self.cascader_error = True
            return {'status': FAILED}
        params = step.get('params')
        step_data = parse_api_case_steps([params['case_related'][-1]], is_step=True)
        prefix_label += step['step_name'] + '-'
        res_status, step_data = run_step_groups(self, step_data, prefix_label, cascader_level, i)
        if cascader_level == 1 and self.cascader_error:
            self.cascader_error = False
            return {'status': FAILED, 'results': '步骤死循环或主计划步骤嵌套的子用例超过10层！'}
        return {'status': res_status, 'results': step_data}

    def var(self, step, prefix_label='', i=0):
        """
          执行类型为全局变量的步骤
        """

        params = step['params']
        if self.envir + '_mode' in params and self.envir + '_source' in params:
            mode, data = params[self.envir + '_mode'], params[self.envir + '_source']
            var = self.parse_source_params(data, mode, i, params_type=API_VAR)
            for name in var.keys():
                self.params_source[VAR_PARAM][name] = {
                    'name': name, 'value': var[name], 'step_name': prefix_label + step['step_name'], 'type': VAR_PARAM,
                    'param_type_id': PY_TO_CONF_TYPE.get(str(type(var[name])), STRING), **self.base_params_source}

    def sql(self, step, prefix_label='', i=0):
        """
          执行类型为SQL的步骤
        """
        params = step.get('params') or step
        db_data = get_proj_envir_db_data(params['sql_proj_related'], envir=self.envir)
        sql = params['sql']
        if db_data:
            if 'database' in params:
                db_data['db_database'] = params['database']
            res = db_connect(db_data)
            if res['status'] == SUCCESS:
                try:
                    sql = parse_param_value(sql, self.default_var, i)
                except DiyBaseException as e:
                    return {'status': FAILED, 'results': f'执行出错：{e}', 'sql': sql}
                sql_res = execute_sql_func(res['db_con'], sql, db_data['db_type'])
                close_db_con(res)
                if sql_res['status'] == SUCCESS:
                    if sql_var := params.get('sql_var'):
                        sql_data = sql_res['data']['sql_data']
                        self.default_var[sql_var] = sql_data
                        self.params_source[VAR_PARAM][sql_var] = {
                            'name': sql_var, 'value': json_dumps(sql_data, JSONEncoder), 'type': VAR_PARAM,
                            'step_name': prefix_label + step['step_name'], 'param_type_id': OBJECT,
                            **self.base_params_source}
                    return sql_res
                return sql_res
            return res
        return {'status': FAILED, 'results': '无效的连接！'}

    def foreach(self, step, prefix_label='', cascader_level=1, i=0):
        """
        循环控制器
        """
        if cascader_level > 15:
            self.cascader_error = True
            return {'status': FAILED}
        params = step['params']
        times_value, break_code = params['times'], params.get('break_code')
        if times_value.isdigit():
            for_times = int(times_value) if int(times_value) <= 999 else 999
        else:  # foreach_times可能为可迭代对象（列表）的情况
            for_times = parse_param_value(times_value, self.default_var)
            if isinstance(for_times, list):
                times = len(for_times)
                for_times = times if times <= 100000 else 100000  # 最多循环10万次
            elif for_times not in ('true', True) and not isinstance(for_times, int):
                raise DiyBaseException(f'无效的循环次数值：{for_times}！')
        if 'steps' not in params:
            steps = set_foreach_tree(ApiForeachStep.objects.filter(step_id=step['id']).values().order_by('id'))
        else:  # 调试时，不会传递foreach_id
            steps = params['steps']
        prefix_label += step['step_name'] + '-'
        res_status, res_data = SUCCESS, []
        if type(for_times) != int and for_times in ('true', True):
            while True:
                # 满足break条件的话则中止循环
                if self.status == INTERRUPT or break_code and run_params_code(
                        break_code, copy.deepcopy(self.default_var), i):
                    break
                run_status, step_data = run_step_groups(
                    self, copy.deepcopy(steps), prefix_label, cascader_level=cascader_level, i=i)
                i += 1
                res_data.append(step_data)
                if run_status == FAILED:
                    res_status = FAILED
        else:
            for _ in range(for_times):
                # 满足break条件的话则中止循环
                if self.status == INTERRUPT or break_code and run_params_code(
                        break_code, copy.deepcopy(self.default_var), i):
                    break
                run_status, step_data = run_step_groups(
                    self, copy.deepcopy(steps), prefix_label, cascader_level=cascader_level, i=i)
                i += 1
                res_data.append(step_data)
                if run_status == FAILED:
                    res_status = FAILED
        if cascader_level == 1:
            if self.cascader_error:
                self.cascader_error = False
                return {'status': FAILED, 'results': '步骤死循环或主计划步骤嵌套的子用例超过15层！'}
        return {'status': res_status, 'results': res_data}

    def parse_source_params(self, data, mode=TABLE_MODE, i=0, params_type='', file_list=None):
        """
        解析请求数据
        """
        res = {}
        if mode == TABLE_MODE:
            for param in data:
                p_name, p_v = param['name'], param.get('value')
                parm_type = param.get('type', {}).get('type') or STRING
                p_name = parse_param_value(p_name, self.default_var, i)
                p_v = parse_param_value(p_v, self.default_var, i)
                res[p_name] = format_parm_type_v(p_v, parm_type)
                if params_type == API_VAR:
                    self.default_var[p_name] = res[p_name]
        elif mode == JSON_MODE:
            if isinstance(data, dict):
                for p_name in data.keys():
                    p_name = parse_param_value(p_name, self.default_var, i)
                    p_v = parse_param_value(data[p_name], self.default_var, i)
                    res[p_name] = p_v
                    if params_type == API_VAR:
                        self.default_var[p_name] = res[p_name]
            else:
                res = parse_param_value(data, self.default_var, i)
        elif mode == TEXT_MODE:
            res = parse_param_value(data, self.default_var, i)
        elif mode == CODE_MODE:
            res = run_params_code(data, copy.deepcopy(self.default_var), i)
            if isinstance(res, dict):
                if params_type == API_VAR:  # 步骤类型为全局变量的话，则将其加入到全局变量中
                    self.default_var.update(res)
            else:
                raise DiyBaseException('返回格式不正确！需要返回一个字典')
        elif mode == FORM_MODE:
            req_data, body_log = {}, {}
            for v in data:
                parm_name, parm_v = v['name'], v['value']
                print('pp', parm_v)
                if isinstance(parm_v, dict) and 'type' in parm_v:
                    if parm_v['type'] == FORM_FILE_TYPE:
                        file_name, file_url = parm_v['name'], parm_v['value']
                        r = requests.get(file_url)
                        if r.status_code == 404:
                            print('ff', file_url)
                            raise NotFoundFileError('未找到上传的文件：' + file_name)
                        with open(file_name, 'wb') as f:
                            f.write(r.content)
                        file_type = os.path.splitext(file_name)[-1]
                        if file_type in ['.xlsx'] and '有变量' in file_name:
                            try:
                                self.parse_excel_var_params(file_name)
                            except Exception as e:
                                file_list.append({'name': file_name, 'file': None})
                                raise Exception('导入报错：' + str(e))
                        file = open(file_name, 'rb')
                        req_data[parm_name] = (file_name, file)
                        body_log[parm_name] = file_name
                        file_list.append({'name': file_name, 'file': file})
                    else:
                        parm_v = parse_param_value(parm_v['value'], self.default_var)
                        body_log[parm_name] = parm_v
                        req_data[parm_name] = (None, parm_v)
                else:
                    body_log = '不是有效的form-data参数！'
                    break
            res = [req_data, body_log]
        return res


def save_results(step_data, case_data):
    """
    执行完成后，写入结果
    """

    ApiCase.objects.bulk_update(case_data, fields=('status', 'report_data', 'latest_run_time'))
    ApiCaseStep.objects.bulk_update(step_data, fields=('status', 'results', 'params'))


def run_step_groups(actuator_obj, step_data, prefix_label='', cascader_level=0, i=0):
    """
    执行步骤合集
    """
    # 默认测试是通过的
    run_status = SUCCESS
    for step in step_data:
        s_type = step['type']
        if step.get('enabled'):
            params = {'actuator_obj': actuator_obj, 's_type': s_type, 'step': step, 'prefix_label': prefix_label,
                      'i': i}
            if s_type in (API_CASE, API_FOREACH):
                params['cascader_level'] = cascader_level + 1
            res = go_step(**params)
            step.update(res)
        else:
            step['status'] = DISABLED
        # step.update({'status': res['status'], 'results': res.get('results')})
        # 当测试计划状态为通过且步骤状态为失败时，就将计划状态改为失败
        if run_status != FAILED and step['status'] == FAILED:
            run_status = FAILED
    return run_status, step_data


def go_step(actuator_obj, s_type, step, i=0, prefix_label='', **extra_params):
    # 执行状态为中断时则直接返回跳过，但下面的处理方式会导致循环器/引用计划直接中断，它们里面的步骤状态不会改变，还是上次的执行结果
    if actuator_obj.status in (INTERRUPT, FAILED_STOP):
        return {'status': SKIP, 'results': '执行被中断！' if s_type not in (API_CASE, API_FOREACH) else None}
    params = {'step': step, 'i': i, 'prefix_label': prefix_label, **extra_params}
    controller_data = step.get('controller_data') or {}
    # 为了避免失败跳过执行出现BUG，plan和foreach不允许设置重试，设置的话会默认不重试
    retry_times = controller_data.get('re_times', 0) if step['type'] not in (API_CASE, API_FOREACH) else 0
    retry_interval, execute_on = controller_data.get('re_interval', 0), controller_data.get('execute_on', '')
    sleep_time = controller_data.get('sleep')
    res = {'status': SUCCESS, 'results': ''}
    if execute_on:
        try:
            res = run_params_code(execute_on, copy.deepcopy(actuator_obj.default_var), i)
            if not res:
                return {'status': SKIP, 'results': '【控制器】执行条件不满足！'}
        except Exception as e:
            if actuator_obj.failed_stop:
                actuator_obj.running_status = INTERRUPT
            return {'status': FAILED, 'results': '【控制器】' + str(e)}
    sleep_time and time.sleep(sleep_time)
    for j in range(retry_times + 1):
        print(step.get('step_name', '') + '，执行次数：' + str(j))
        try:
            res = getattr(actuator_obj, s_type)(**params) or {'status': SUCCESS}
        except Exception as e:  # 捕获步骤执行过程的异常
            res = {'status': FAILED, 'results': str(e)}
        if res['status'] == FAILED:
            if j < retry_times:
                time.sleep(retry_interval)
        else:
            break
    res['retried_times'] = j
    # 只有成功和跳过执行的步骤才不记录日志，失败和中断的还是会记录日志
    if actuator_obj.only_failed_log and res['status'] in (SUCCESS, SKIP) and s_type != API_CASE:
        return {'status': res['status'], 'retried_times': res['retried_times']}
    if res['status'] == FAILED:
        if actuator_obj.failed_stop:
            actuator_obj.status = FAILED_STOP
    # if res['status'] == SPEND_TIME_OUT:
    #     res['status'] = FAILED
    return res


def monitor_interrupt(user_id, actuator_obj):
    while True:
        time.sleep(3)
        print('monitor_interrupt', actuator_obj.status)
        exec_status = UserCfg.objects.filter(user_id=user_id).values_list('exec_status', flat=True).first()
        if exec_status == INTERRUPT:
            actuator_obj.status = INTERRUPT
            quit()
        elif exec_status == WAITING:
            quit()


def run_api_case_func(case_data, user_id, cfg_data=None, temp_params=None):
    """
    执行api用例的主方法
    执行测试计划：case_data={case_id:[step1,step2,step3]}
    实时调试/步骤中计划：case_data=[step1,step2,step3]
    temp_params为空的话则查询用户的参数来测试。
    """

    res_step_objs, res_case_objs = [], []
    actuator_obj = ApiCasesActuator(user_id, cfg_data=cfg_data, temp_params=temp_params)
    thread = MyThread(target=monitor_interrupt, args=[user_id, actuator_obj])
    thread.start()
    if isinstance(case_data, dict):
        for case_id in case_data.keys():
            start_time = datetime.datetime.now()
            case_objs = ApiCase.objects.filter(id=case_id).first()
            if case_objs:
                case_objs.status = RUNNING
                case_objs.save(update_fields=['status'])
            report_dict = {'envir': actuator_obj.envir, 'start_time': start_time.strftime('%Y-%m-%d %H:%M:%S'),
                           'steps': []}
            actuator_obj.base_params_source['case_id'] = case_id
            # 默认测试是通过的
            case_status, step_data = run_step_groups(actuator_obj, case_data[case_id])
            report_dict['steps'] = step_data
            for step in step_data:
                res_step_objs.append(ApiCaseStep(**step))
            end_time = datetime.datetime.now()
            report_dict['spend_time'] = format((end_time - start_time).total_seconds(), '.1f')
            if actuator_obj.status in (INTERRUPT, FAILED_STOP):
                case_status = actuator_obj.status
            res_case_objs.append(
                ApiCase(id=case_id, status=case_status, latest_run_time=end_time, report_data=report_dict))
        save_results(res_step_objs, res_case_objs)
        return {'params_source': actuator_obj.params_source}


def parse_api_case_steps(case_ids=None, is_step=False):
    """
    转化API计划步骤
    is_step:false代表非步骤中的用例，即外层计划列表中选中执行的用例
    """
    step_data = []
    if case_ids:
        step_data = list(ApiCaseStep.objects.filter(case_id__in=case_ids).annotate(
        ).select_related('case', 'case__module', 'api').values(
            'id', 'step_name', 'type', 'case_id', 'status', 'params', 'results', 'api_id',
            'controller_data', 'enabled').order_by('id'))
        if not is_step:  # 如果非测试计划步骤而是执行测试用例，需要转为{case_id:[step,step],case_id2:[step,step]}的形式
            case_data = {case_id: [] for case_id in case_ids}  # {case1:steps,case2:steps}
            for step in step_data:
                case_data[step['case_id']].append(step)
            return case_data
    return step_data


def parse_create_foreach_steps(save_step_objs, foreach_step, step_id, next_id, parent_id=None):
    """
    格式化循环控制器步骤为创建数据
    """

    for step in foreach_step:
        step.pop('results', None)
        if (s_type := step['type']) == API:
            step['api_id'] = step['params']['api_id']
        elif s_type == API_CASE:
            step['quote_case_id'] = step['params']['case_related'][-1]
        step.update({'step_id': step_id, 'id': next_id, 'parent_id': parent_id})
        save_step_objs.append(ApiForeachStep(**step))
        next_id += 1
        if s_type == API_FOREACH:
            parse_create_foreach_steps(save_step_objs, step['params'].pop('steps', []), step_id, next_id,
                                       step['id'])
    return next_id


def set_foreach_tree(_list):
    """
    生成循环控制器树
    """
    _dict, tree = {}, []
    for i in _list:
        _dict[i['id']] = i
        if i['type'] == API_FOREACH:
            i['params']['steps'] = []
    for i in _list:
        node = i
        if node['parent_id'] is not None:
            _dict[node['parent_id']]['params']['steps'].append(node)
        else:
            tree.append(node)
    return tree


def copy_cases_func(request, case_model, step_model, foreach_step_model=None):
    """
    复制用例方法
    """

    req_data = request.data
    case_obj = case_model.objects.get(id=req_data['case_id'])
    case_obj.creater_id = request.user.id
    case_obj.id = None
    case_obj.status = WAITING
    case_obj.name = case_obj.name + '-' + str(int(time.time()))
    case_steps = step_model.objects.filter(case_id=req_data['case_id']).values()
    case_obj.save()
    step_objs = []
    foreach_steps_obj = []
    next_id = (ApiCaseStep.objects.aggregate(Max('id')).get('id__max') or 0) + 1
    for step in case_steps:
        step['case_id'] = case_obj.id
        old_step_id = step.pop('id')
        step['id'] = next_id
        step.pop('results', None)
        step_objs.append(step_model(**step))
        print('ada', step)
        if step['type'] == API_FOREACH:
            for for_step in ApiForeachStep.objects.filter(step_id=old_step_id).values():
                for_step.pop('id')
                for_step['step_id'] = next_id
                print('fa', for_step)
                foreach_steps_obj.append(ApiForeachStep(**for_step))
        next_id += 1
    step_model.objects.bulk_create(step_objs)
    if foreach_steps_obj:
        ApiForeachStep.objects.bulk_create(foreach_steps_obj)
    return Response(data={'msg': "复制成功！"})
